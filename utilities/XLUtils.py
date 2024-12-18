import openpyxl
from openpyxl.styles import PatternFill
# global workbook
# global sheet

def load_excel(filepath, sheetName):
    global workbook
    global sheet
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetName]
    
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)  
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

# def get_cell_data(rownum, columnno):
#     return sheet.cell(rownum, columnno).value

def get_data_as_list_tuples():
    sheet_cells=[]
    for i in range(1, sheet.max_row):
        row_cells=[]
        for cell in sheet[i+1]:
            row_cells.append(cell.value)
        sheet_cells.append(tuple(row_cells))
    return sheet_cells

def fillGreenColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                       end_color='60b212',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

def fillRedColor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                       end_color='ff0000',
                       fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)